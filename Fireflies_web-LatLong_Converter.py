import openpyxl
import time
import re
from selenium import webdriver
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC

#File name
file_name = 'data.xlsx'

#Other
dead_lat = '40.7648'
dead_lng = '-73.9808'

#Site URLs
lat_lng_url = 'https://gps-coordinates.org/my-location.php?lat={}&lng={}'
gps_reading_address_url = 'https://gps-coordinates.org/coordinate-converter.php' #Use VPN, site complains after a while
google_maps_search_url = 'https://www.google.com/maps/search/{}'


#Define workbook
workbook = openpyxl.load_workbook(file_name)

# Open a web browser
driver = webdriver.Chrome(ChromeDriverManager().install())

# Open the Excel file and sheet
workbook = openpyxl.load_workbook(file_name)
# Select the sheet
sheet = workbook.active

wait = WebDriverWait(driver, 3)

#Regex checks if either a decimal-based lat/lng pair or a DMS-based lat/lng pair, with/out whitespace or commas
decimal_pattern = r'^(-?\d+(.\d+)?),\s*(-?\d+(.\d+)?)$'
dms_pattern = r'^(\d+)°(\d+)\'(\d+(\.\d+)?)"([NEWS])\s(\d+)°(\d+)\'(\d+(\.\d+)?)"([NEWS])$'

def check_lat_lng(cell):

    decimal_match = re.search(decimal_pattern, cell)
    dms_match = re.search(dms_pattern, cell)

    if decimal_match:
        lat = decimal_match.group(1)
        lng = decimal_match.group(3)
        #Debug
        print(f'Found Lat/Long in cell: {cell}')
        sheet.cell(row=i, column=7).value = "<100 ft."
        return lat, lng
    elif dms_match:
        lat_degree = float(dms_match.group(1))
        lat_minute = float(dms_match.group(3))
        lat_second = float(dms_match.group(4))
        lat_direction = dms_match.group(5)

        lng_degree = float(dms_match.group(6))
        lng_minute = float(dms_match.group(8))
        lng_second = float(dms_match.group(9))
        lng_direction = dms_match.group(10)

        lat = lat_degree + lat_minute / 60 + lat_second / 3600
        lng = lng_degree + lng_minute / 60 + lng_second / 3600
        if lat_direction in ['S', 'W']:
            lat = -lat
        if lng_direction in ['S', 'W']:
            lng = -lng
        #Debug
        print(f'Converted DMS to Lat/Long in cell: {cell}')
        sheet.cell(row=i, column=7).value = "<100 ft."
        return lat, lng
    else:
        try:
            print(f'Reading written street address in cell: {cell}')
            
            return read_street_address(cell)            

        except Exception:
            #Debug
            print(f'Skipping cell {cell}')
            return None, None


def read_street_address(cell):

    cell_city = str((sheet.cell(row=i, column=2)).value)
    written_address = cell + " " + cell_city #+ " " + "Utah"  # Helps to add this on there if the address is totally busted

    try:
        #Try GPs-coords.org
        driver.get(gps_reading_address_url)

        #Write address to search field
        address_field = wait.until(EC.presence_of_element_located((By.ID, "address")))
        address_field.clear()
        address_field.send_keys(written_address)

        #Convert
        submit_button = driver.find_element(By.ID, "btnGetGpsCoordinates")
        submit_button.click()

        #Find matched Lat/Lng
        time.sleep(1) #Let lat/lng load
        lat = wait.until(EC.presence_of_element_located((By.ID, "latitude"))).get_attribute("value")
        lng = wait.until(EC.presence_of_element_located((By.ID, "longitude"))).get_attribute("value")

        #Will keep the default lat/lng of NYC if the address didn't take the page anywhere, check for those defaults
        if lat == dead_lat or lng == dead_lng:
            #Leave this site and go try Google Maps; larger search results, more variance
            raise Exception("GPS-Coord.org couldn't find address")
        
        print('Includes slight variance')
        sheet.cell(row=i, column=7).value = "<0.5 mi."
        return lat, lng


    except Exception:

        #check google maps if GPS-coords cant find it
        #If you can find DkEaL, read the lat long that is in the url bar
        #Side note, google seems to dynamically create their div id's? I guess? DkEaL seems to be the one created for this webdriver though, who knows
        #sometimes also Io6YTe
        try:
            driver.get(google_maps_search_url.format(written_address))

            # The url updates from the search querey to the formatted query with other metadata, including the lat/lng we need
            time.sleep(3)

            #Checking for the presence of an autocompleted address, not going to use it, but it means that the lat/lng is at least sort of close
            try:
                # Check for class name 'DkEaL'
                google_maps_address = wait.until(EC.presence_of_element_located((By.CLASS_NAME, 'DkEaL')))
            except TimeoutError:
                # If class name 'DkEaL' is not found, check for class name 'Io6YTe'
                try:
                    google_maps_address = wait.until(EC.presence_of_element_located((By.CLASS_NAME, 'Io6YTe')))
                except TimeoutError:
                    # If class name 'Io6YTe' is not found, raise an exception
                    raise Exception("Could not find element with class name 'DkEaL' or 'Io6YTe'")
                
            current_url = driver.current_url

            # Extract the latitude and longitude from the URL
            lat_long_pattern = re.compile(r'@(-?\d+\.\d+),(-?\d+\.\d+)')
            match = lat_long_pattern.search(current_url)
            lat, lng = match.groups()

            print('Includes high variance')
            sheet.cell(row=i, column=7).value = "0.6-19 mi."

            return lat, lng

        except Exception:
            try:
                # If there is not a close enough address match, grab the url and pull the lat/lng of the map preview
                # It is centered around the "suggested alternatives", use that lat/lng as the point for GPS-coords.org

                current_url = driver.current_url
                
                # Extract the latitude and longitude from the URL
                lat_long_pattern = re.compile(r'@(-?\d+\.\d+),(-?\d+\.\d+)')
                match = lat_long_pattern.search(current_url)
                lat, lng = match.groups()

                print('Includes extreme variance')
                sheet.cell(row=i, column=7).value = ">20 mi."

                return lat, lng
                
            except Exception:
                #If you get here then the input was so far off that you're just hosed
                print("Google maps couldn't find anything")
                sheet.cell(row=i, column=7).value = "Cannot find an address"
                print(f'Skipping cell {cell}')          
                return None, None


def convert_matched_address(lat, lng):

    # Providing URL
    url = lat_lng_url.format(lat, lng)

    # Navigate to the page
    driver.get(url)
    #Wait for page to load
    time.sleep(1)

    # Find the input element with the specified ID and extract value
    address_field = wait.until(EC.presence_of_element_located((By.ID, "address"))).get_attribute("value")

    # Use a dictionary to map abbreviations to state names
    states = {
    "AZ": "Arizona",
    "CA": "California",
    "CO": "Colorado",
    "ID": "Idaho",
    "MT": "Montana",
    "NV": "Nevada",
    "NM": "New Mexico",
    "OR": "Oregon",
    "UT": "Utah",
    "WA": "Washington",
    "WY": "Wyoming",
    "Error": "Couldn't find state"
    }

    #Format state info
    state_abbrv = re.search(r",\s([A-Z]{2})", address_field)
    if state_abbrv:
        state_abbrv = state_abbrv.group(1)
    else:
        state_abbrv = "Error"
        
    state_FullName = states.get(state_abbrv, "Error")
    
    # Save the lat, long, formatted address, state to the spreadsheet
    sheet.cell(row=i, column=3).value = float(lat)
    sheet.cell(row=i, column=4).value = float(lng)
    sheet.cell(row=i, column=5).value = address_field
    sheet.cell(row=i, column=6).value = state_FullName

    #Debug
    print(f'Wrote matched address to cell: {cell}')


def check_for_duplicates():
    print("Checking for duplicat entries now...")
    for i in range(1, sheet.max_row + 1):
        # Check if the value in the second row is a duplicate
        if sheet.cell(row=i, column=3).value == sheet.cell(row=i+1, column=3).value:
            sheet.cell(row=i+1, column=8).value = "Possible Duplicate Entry"
    
    workbook.save(file_name)


# --- # MAIN # --- #


# Iterate through the cells in the selected column
for i in range(2, sheet.max_row + 1):
    # Check if the row has empty cells in the next four columns
    if (sheet.cell(row=i, column=3).value and sheet.cell(row=i, column=4).value 
        and sheet.cell(row=i, column=5).value and sheet.cell(row=i, column=6).value and sheet.cell(row=i, column=7).value):
        print(f'Already populated, skipping cell: {i}')
        continue

    # Next Cell
    cell = sheet.cell(row=i, column=1)
    # Parse cell contents
    lat, lng = check_lat_lng(str(cell.value))
    # If returns not "None, None", covnert the parsed lat/lng to a formatted street address and write it to the sheet
    if lat and lng:
        convert_matched_address(lat, lng)
        workbook.save(file_name)


# Check for double entries and mark them
check_for_duplicates()


#Clean up
print(f'Save and close {file_name}')
workbook.save(file_name)
workbook.close()
driver.close()

